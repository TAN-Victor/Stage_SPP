###===========================================================================================
###==================================training.py==============================================
###===========================================================================================
### This file contains the training functions that are used in Reinforcement Learning for
### training the model (SSP-MV).
###===========================================================================================


#=============================================================================================
#===================================Importing Libraries=======================================
#=============================================================================================
from time import time
import torch
import numpy as np
import pandas as pd
from environment import Environment
import matplotlib.pyplot as plt
import gurobipy as gp
import scipy.linalg as la
from stable_baselines3.common.vec_env import DummyVecEnv
from stable_baselines3 import A2C, PPO
import openpyxl
from tqdm import tqdm
from stable_baselines3.common.callbacks import EvalCallback, StopTrainingOnNoModelImprovement

np.set_printoptions(threshold=np.inf)
#=============================================================================================


#=============================================================================================
#==================================Default Variables==========================================
#=============================================================================================
Default_model_dict = {
    "model_name": (A2C, type),
    "policy": ("MlpPolicy", str),
    "device": ("cuda:0" if torch.cuda.is_available() else "cpu", str),
    "seed": (5, int),
}

Default_training_dict = {
    "max_iteration": (5, int),
    "n_timesteps": (10000, int),
}

Default_env_dict = {
    "data": (None, pd.DataFrame),
    "sigma": (0.1, float),
    "window_size": (-1, int),
    "random_nn": (False, bool),
    "random_data": (False, bool),
    "cardinality_constraint_mode": ("None", str),
    "cardinality_constraint": (None, int),
    "shrinkage": (True, bool),
    "lambda_regularization": (0, float), 
}

Default_policy_dict = {
    "net_arch": ([8, 8], list),
}

Default_vector_dict = {
    "vector_size": (5, int),
    "seed": (5, int),
    "subsample": (-1, int),
    "shuffle": (True, bool),
}
#=============================================================================================


#=============================================================================================
#=====================================Functions===============================================
#=============================================================================================
def get_mu_sigma(data, shrinkage = True):
    """
    Get the (shrinked) expected return and the (shrinked) covariance given the data.
    Shrinkage method is used to estimate the expected return and the covariance and reduce error and noise.
    Shrinkage method is based on the work of Jorion (1986).
    
    Attributes:
        data: the data to use for the shrinkage method
        shrinkage: the shrinkage

    Returns:
        np.ndarray: The expected return.
        np.ndarray: The covariance.
    """
    mu = data.mean()
    sigma = data.cov()
    if shrinkage:
        m = len(data)
        n = len(data.columns)
        try:
            sigma_s = (m - 1) / (m - n - 2) * sigma
            z = la.solve(sigma_s, np.ones(n), assume_a='pos')
            mu_0 = mu @ z / z.sum()
            d = mu - mu_0 * np.ones(n)
            y = la.solve(sigma_s, d, assume_a='pos')
            w = (n + 2) / (n + 2 + m * d @ y)
            mu_s =  (1 - w) * mu + w * mu_0 * np.ones(n)
        except:
            mu_s = mu
            sigma_s = sigma
        return mu_s, sigma_s
    else:
        return mu, sigma



def DRL_vect(vector_dict, model_dict, training_dict, env_dict, policy_dict):
    """
    Trains and tests the model for the given number of iterations. Vectorized version.

    Attributes:
        vector_dict (dict): the parameters for the vectorized version
            - vector_size (int): the number of environments to run in parallel
            - seed (int): the seed to use for shuffling the columns of the data, default is max_iteration
            - subsample (int): the number of columns to subsample from the data, default is -1 which means all columns
            - shuffle (bool): whether to shuffle the columns of the data across the environments, default is True, can't be False if subsample is not -1
        model_dict (dict): the model to use and its parameters for learning
            - model_name (class from stable_baselines3): the model to use
            - policy (str): the policy to use, str class from stable_baselines3
            - device (str): the device to use, default is "cuda:0"
            - seed (int): the seed to use for the model, default is max_iteration
            - n_steps (int): the number of steps to use for training before updating the model, default is 5 for A2C, 2048 for PPO
            - batch_size (int) [only for PPO]: the batch size to use for training, default is 64
        training_dict (dict): the training parameters
            - max_iteration (int): the number of iterations to train and test the model on
            - n_timesteps (int): the total timesteps to train the model on
            - save_model (str): the name of the file to save the model to
        env_dict (dict): the environment parameters
            - sigma (float): the sigma, as the trade-off risk, default is 0.1
            - window_size (int): the window size, default is -1
            - random_nn (bool): whether to randomize the columns of the data, which would randomize the entries of the neural network, default is False
            - random_data (bool): whether to randomize the rows of the data, default is False
            - cardinality_constraint_mode (str): the cardinality constraint mode, default is "None"
            - cardinality_constraint (int): the cardinality constraint
            - shrinkage (bool): whether to use shrinkage, default is True
            - lambda_regularization (float): the lambda regularization for Ridge regularization, default is 0
        policy_dict (dict): the policy parameters
            - net_arch (list): the architecture of the neural network, default is [8, 8]
    """
    m, n = env_dict["data"].shape
    info = {"training time": 0, "testing time": 0}
    
    time_start = time()
    env_list = []
    env_data_copy = env_dict["data"].copy()
    def env_creator(seed):
        np.random.seed(seed)
        if vector_dict["shuffle"] or vector_dict["subsample"] != -1:
            random_index = np.random.choice(n, vector_dict["subsample"], replace=False)
            data_tmp = env_data_copy.iloc[:, random_index]
            env_dict["data"] = data_tmp
        env = Environment(**env_dict)
        return env
    
    for i in range(vector_dict["vector_size"]):
        env_list.append(env_creator(i))
    DummyVec = DummyVecEnv([(lambda env: lambda: env)(outer_env) for outer_env in env_list])
    stop_train_callback = StopTrainingOnNoModelImprovement(max_no_improvement_evals=5, min_evals=1000)
    eval_callback = EvalCallback(DummyVec, eval_freq= 1000, callback_after_eval=stop_train_callback, warn=False, verbose=0)
    model_name = model_dict["model_name"]
    model_dict["env"] = DummyVec
    model_dict["policy_kwargs"] = policy_dict
    model_dict["seed"] = training_dict["max_iteration"]
    model_dict_copy = model_dict.copy()
    del model_dict_copy["model_name"]
    model = model_name(**model_dict_copy)
    model.learn(total_timesteps = training_dict["n_timesteps"], progress_bar = True, callback=eval_callback)
    time_end = time()
    print(f"Training time: {round(time_end - time_start, 3)} seconds.")
    if "save_model" in training_dict:
        model.save(training_dict["save_model"])
    else:
        model.save(f"models/{type(model).__name__}_vectorized{vector_dict['vector_size']}_ws{env_dict['window_size']}_rnn{env_dict['random_nn']}_rd{env_dict['random_data']}_ccm{env_dict['cardinality_constraint_mode']}_cc{env_dict['cardinality_constraint']}_sh{env_dict['shrinkage']}_lr{env_dict['lambda_regularization']}_netarch{'_'.join(str(x) for x in policy_dict['net_arch'])}.zip")
    info["training time"] += time_end - time_start

    time_start = time()
    env_list = []
    for i in range(vector_dict["vector_size"]):
        env_list.append(env_creator(training_dict["max_iteration"]))
    DummyVec = DummyVecEnv([(lambda env: lambda: env)(outer_env) for outer_env in env_list])
    obs = DummyVec.reset()
    model.set_random_seed(training_dict["max_iteration"])
    for _ in tqdm(range(m - 1)):
        action, states = model.predict(obs)
        obs, rewards, done, info_d = DummyVec.step(action)
    memory_expected_return = DummyVec.env_method(method_name="get_memory_expected_return")
    memory_variance = DummyVec.env_method(method_name="get_memory_variance")
    memory_weights = DummyVec.env_method(method_name="get_memory_weights")
    expected_return = np.array(memory_expected_return)
    variance = np.array(memory_variance)
    weights = np.array(memory_weights)
    time_end = time()
    info["testing time"] += time_end - time_start

    return expected_return, variance, weights, info



def plot_results(expected_return, variance, data, figname):
    """
    Plots the expected return and variance of the model.
    
    Attributes:
        expected_return: the expected return of the model
        variance: the variance of the model
        data: the data to plot
    """
    fig, ax = plt.subplots(2, 1, figsize=(17, 10))
    ax[0].plot(data.index, expected_return.mean(axis=0), label="Expected Return")
    ax[0].fill_between(data.index, expected_return.mean(axis=0) - expected_return.std(axis=0), expected_return.mean(axis=0) + expected_return.std(axis=0), color='b', alpha=0.2)
    ax[0].set_title("Expected Return: Mean = {:.2f}, Std = {:.2f}".format(expected_return.mean(), expected_return.std()))
    ax[0].set_xlabel("Date")
    ax[0].set_ylabel("Expected Return")
    ax[0].legend()
    
    ax[1].plot(data.index, variance.mean(axis=0), label="Variance")
    ax[1].fill_between(data.index, variance.mean(axis=0) - variance.std(axis=0), variance.mean(axis=0) + variance.std(axis=0), color='r', alpha=0.2)
    ax[1].set_title("Variance: Mean = {:.2f}, Std = {:.2f}".format(variance.mean(), variance.std()))
    ax[1].set_xlabel("Date")
    ax[1].set_ylabel("Variance")
    ax[1].legend()
    
    plt.savefig(figname)
    plt.show()
    return None



def print_results(expected_return, variance, weights, model_dict, env_dict, info):
    """
    Prints the results of the model.
    
    Attributes:
        expected_return: the expected return of the model
        variance: the variance of the model
        weights: the weights of the model
        info (dict): the information about the training and testing times
    """
    avg_expected_return = expected_return.mean(axis=0)
    avg_variance = variance.mean(axis=0)
    optimal_values = env_dict["sigma"] * expected_return[:, -1] - (1 - env_dict["sigma"]) * variance[:, -1]
    max_index = np.argmax(optimal_values, axis=0)
    m, n = expected_return.shape
    max_weights = weights[max_index, -1, :]
    info["optimal_value"] = (env_dict["sigma"] * avg_expected_return[-1] - (1 - env_dict["sigma"]) * avg_variance[-1]).round(4)
    info["optimal_weights"] = max_weights.round(4)
    info["expected_return"] = avg_expected_return[-1].round(4)
    info["variance"] = avg_variance[-1].round(4)
    print("### Results for {}".format(model_dict["model_name"]))
    print(" | Optimal value: ", info["optimal_value"])
    print(" | Maximum Value weights: ", info["optimal_weights"])
    print(" | Expected return: ", info["expected_return"])
    print(" | Variance: ", info["variance"])
    print(" | Training time: ", info["training time"])
    print(" | Testing time: ", info["testing time"])
    return info



def print_gurobi(data, sig, filename:str, constraint: int = None, name_dataset: str = None, print_results: bool = True, verbose: bool = False):
    """
    Prints the results of the Gurobi model.
    
    Attributes:
        data: the data
        sig: the sigma for the trade-off risk
        filename: the name of the Excel file to save the results to
        name_dataset: the name of the dataset
        constraint: the cardinality constraint
        print_results: whether to print the results
        verbose: whether to print the results
    """
    m, n = data.shape
    mu, sigma = data.mean(), data.cov()

    model_gurobi = gp.Model()
    if not verbose:
        model_gurobi.Params.LogToConsole = 0
    model_gurobi.setParam('TimeLimit', 1800)    # 30 minutes
    model_gurobi.setParam('Threads', 1)
    x = model_gurobi.addMVar(len(mu), lb=0, ub=1, name="x")

    model_gurobi.addConstr(x.sum() == 1, "Budget_Constraint")
    if constraint is not None:
        b = model_gurobi.addMVar(len(mu), vtype=gp.GRB.BINARY, name="b")
        model_gurobi.addConstr(x <= b, name = "Indicator")
        model_gurobi.addConstr(x >= b * 0.01, name = "Minimum_Indicator")
        model_gurobi.addConstr(b.sum() <= constraint, name = "Cardinality_Constraint")

    model_gurobi.setObjective(sig * (mu.to_numpy() @ x) - (1-sig) * (x @ sigma.to_numpy() @ x), gp.GRB.MAXIMIZE)
    model_gurobi.optimize()

    if print_results:
        print("### Results for Gurobi")
        print(" | Optimal value: ", model_gurobi.objVal)
        print(" | Optimal weights: ", x.X.round(6))
        print(" | Expected return: ", mu @ x.X)
        print(" | Variance: ", x.X @ sigma.to_numpy() @ x.X)
        print(" | Gap: ", model_gurobi.MIPGap)
        print(" | Solution time: ", model_gurobi.Runtime)
    info = {}
    info["optimal_value"] = model_gurobi.objVal
    info["optimal_weights"] = x.X.round(6)
    info["expected_return"] = mu @ x.X
    info["variance"] = x.X @ sigma.to_numpy() @ x.X
    info["gap"] = model_gurobi.MIPGap
    info["solution_time"] = model_gurobi.Runtime
    info["status"] = model_gurobi.Status
    if filename is not None:
        save_to_xlsx_gurobi(constraint, sig, info, name_dataset, filename)
    return None



def SSP_MV_vect(vector_dict, model_dict, training_dict, env_dict, policy_dict, xlsx_name):
    """
    Checks the dictionary and its keys.
    Trains and tests the model for the given number of iterations.

    Attributes:
        vector_dict (dict): the parameters for the vectorized version
            - vector_size (int): the number of environments to run in parallel
            - seed (int): the seed to use for shuffling the columns of the data, default is max_iteration
            - subsample (int): the number of columns to subsample from the data, default is -1 which means all columns
            - shuffle (bool): whether to shuffle the columns of the data across the environments, default is True, can't be False if subsample is not -1
        model_dict (dict): the model to use and its parameters for learning
            - model_name (class from stable_baselines3): the model to use
            - policy (str): the policy to use, str class from stable_baselines3
            - device (str): the device to use, default is "cuda:0"
            - seed (int): the seed to use for the model, default is max_iteration
            - n_steps (int): the number of steps to use for training before updating the model, default is 5 for A2C, 2048 for PPO
            - batch_size (int) [only for PPO]: the batch size to use for training, default is 64
        training_dict (dict): the training parameters
            - max_iteration (int): the number of iterations to train and test the model on
            - n_timesteps (int): the total timesteps to train the model on
            - save_model (str): the name of the file to save the model to
        env_dict (dict): the environment to use and its parameters
            - data (pandas.DataFrame): the dataset
            - sigma (float): the sigma, as the trade-off risk, default is 0.1
            - window_size (int): the window size, default is -1
            - random_nn (bool): whether to randomize the columns of the data, which would randomize the entries of the neural network, default is False
            - random_data (bool): whether to randomize the rows of the data, default is False
            - cardinality_constraint_mode (str): the cardinality constraint mode, default is "None"
            - cardinality_constraint (int): the cardinality constraint
            - shrinkage (bool): whether to use shrinkage, default is True
            - lambda_regularization (float): the lambda regularization for Ridge regularization, default is 0
        policy_dict (dict): the policy to use and its parameters
            - net_arch (list): the architecture of the neural network, default is [8, 8]
        xlsx_name (str): the name of the Excel file to save the results to
    """
    if env_dict["data"] is None or model_dict is None or training_dict is None or policy_dict is None or env_dict is None or vector_dict is None:
        print("No data provided.")
        return None
    
    model_dict = update_dict_with_default(model_dict, Default_model_dict)
    training_dict = update_dict_with_default(training_dict, Default_training_dict)
    env_dict = update_dict_with_default(env_dict, Default_env_dict)
    policy_dict = update_dict_with_default(policy_dict, Default_policy_dict)
    vector_dict = update_dict_with_default(vector_dict, Default_vector_dict)

    model_dict, training_dict, vector_dict = check_dict(model_dict, training_dict, env_dict, vector_dict)

    expected_return, variance, weights, info = DRL_vect(vector_dict, model_dict, training_dict, env_dict, policy_dict)
    plot_results(expected_return, variance, env_dict["data"].iloc[1:, :], f"images/{model_dict['model_name'].__name__}_ws{env_dict['window_size']}_rnn{env_dict['random_nn']}_rd{env_dict['random_data']}_ccm{env_dict['cardinality_constraint_mode']}_cc{env_dict['cardinality_constraint']}_sh{env_dict['shrinkage']}_lr{env_dict['lambda_regularization']}_netarch{'_'.join(str(x) for x in policy_dict['net_arch'])}.png")
    info_new = print_results(expected_return, variance, weights, model_dict, env_dict, info)
    save_to_xlsx(vector_dict, model_dict, training_dict, env_dict, policy_dict, info_new, xlsx_name)
    return None



def update_dict_with_default(dict, default_dict):
    """
    Updates the dictionary with the default dictionary.
    Also checks for type errors.

    Attributes:
        dict (dict): the dictionary to update
        default_dict (dict): the default dictionary
    """
    message = []
    for key, (default_value, expected_type) in default_dict.items():
        if key not in dict:
            message.append(f"Key {key} not in dict. Using default value {default_value}.")
            dict[key] = default_value
        elif not isinstance(dict[key], expected_type):
            message.append(f"Type error for key {key}. Expected type {expected_type}. Using default value {default_value}.")
            dict[key] = default_value
    if message:
        print(" - ".join(message))
    return dict



def check_dict(model_dict, training_dict, env_dict, vector_dict = None):
    """
    Checks the dictionary for errors.

    Attributes:
        model_dict (dict): the model to use and its parameters for learning
        training_dict (dict): the training parameters
    """
    if model_dict["model_name"] == A2C:
        model_dict.pop("batch_size", None)
        if "n_steps" not in model_dict:
            model_dict["n_steps"] = 5
    elif model_dict["model_name"] == PPO:
        if "batch_size" not in model_dict:
            model_dict["batch_size"] = 64
        if "n_steps" not in model_dict:
            model_dict["n_steps"] = 64
    model_dict["seed"] = training_dict["max_iteration"]
    if env_dict["sigma"] > 1 or env_dict["sigma"] < 0:
        env_dict["sigma"] = 0.1
        print("Sigma must be between 0 and 1. Using default value 0.1.")
    if vector_dict is not None:
        vector_dict["seed"] = training_dict["max_iteration"]
        if vector_dict["subsample"] != -1 and not vector_dict["shuffle"]:
            vector_dict["shuffle"] = True
            print("Can't have shuffle = False if subsample is not -1. Setting shuffle = True.")
        if vector_dict["subsample"] == -1:
            vector_dict["subsample"] = env_dict["data"].shape[1]
        return model_dict, training_dict, vector_dict
    return model_dict, training_dict



def save_to_xlsx(vector_dict, model_dict, training_dict, env_dict, policy_dict, info, xlsx_name):
    """
    Saves the results to an Excel file.

    Attributes:
        vector_dict (dict): the parameters for the vectorized version
            - vector_size (int): the number of environments to run in parallel
            - seed (int): the seed to use for shuffling the columns of the data, default is max_iteration
            - subsample (int): the number of columns to subsample from the data, default is -1 which means all columns
            - shuffle (bool): whether to shuffle the columns of the data across the environments, default is True, can't be False if subsample is not -1
        model_dict (dict): the model to use and its parameters for learning
            - model_name (class from stable_baselines3): the model to use
            - policy (str): the policy to use, str class from stable_baselines3
            - device (str): the device to use, default is "cuda:0"
            - seed (int): the seed to use for the model, default is max_iteration
            - n_steps (int): the number of steps to use for training before updating the model, default is 5 for A2C, 2048 for PPO
            - batch_size (int) [only for PPO]: the batch size to use for training, default is 64
        training_dict (dict): the training parameters
            - max_iteration (int): the number of iterations to train and test the model on
            - n_timesteps (int): the total timesteps to train the model on
            - save_model (str): the name of the file to save the model to
        env_dict (dict): the environment to use and its parameters
            - data (pandas.DataFrame): the dataset
            - sigma (float): the sigma, as the trade-off risk, default is 0.1
            - window_size (int): the window size, default is -1
            - random_nn (bool): whether to randomize the columns of the data, which would randomize the entries of the neural network, default is False
            - random_data (bool): whether to randomize the rows of the data, default is False
            - cardinality_constraint_mode (str): the cardinality constraint mode, default is "None"
            - cardinality_constraint (int): the cardinality constraint
            - shrinkage (bool): whether to use shrinkage, default is True
            - lambda_regularization (float): the lambda regularization for Ridge regularization, default is 0
        policy_dict (dict): the policy to use and its parameters
            - net_arch (list): the architecture of the neural network, default is [8, 8]
        info (dict): the information about the training and testing times, and values
        xlsx_name (str): the name of the Excel file to save the results to
    """
    if "-" in xlsx_name:
        sheet_name = xlsx_name.split("-")[-1]
        xlsx_name = xlsx_name.split("-")[0]

    try:
        workbook = openpyxl.load_workbook(xlsx_name)
    except FileNotFoundError:
        print("File not found. Creating a new one.")
        workbook = openpyxl.Workbook()
    if sheet_name not in workbook.sheetnames:
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(["Model", "Policy", "Vectorized", "Vector Size", "Shuffle Vector", "Max Iteration", "Timesteps", "Data", "n_steps", "batch_size", "net_arch", "random_nn", "random_data", "cardinality_constraint", "card mode", "lambda", "shrinkage", "sigma", "Value", "Return", "Variance", "Time", "Optimal Weights"])
        print("Sheet created.")
    else:
        sheet = workbook[sheet_name]

    def find_first_empty_row(sheet):
        for row in range(2, 100000):
            if sheet[f"A{row}"].value is None:
                return row
        
    def values_exist(sheet, values):
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            if row[:len(values) - 2] == tuple(values[:-2]):
                print("Values exist.")
                return True
        print("Values don't exist. Adding values.")
        return False
    
    values = [model_dict["model_name"].__name__, model_dict["policy"], "Yes", vector_dict["vector_size"], vector_dict["shuffle"], training_dict["max_iteration"], training_dict["n_timesteps"], env_dict["data"].shape[1], model_dict["n_steps"], model_dict["batch_size"] if "batch_size" in model_dict else "-", str(policy_dict["net_arch"]), env_dict["random_nn"], env_dict["random_data"], env_dict["cardinality_constraint"], env_dict["cardinality_constraint_mode"], env_dict["lambda_regularization"], env_dict["shrinkage"], env_dict["sigma"], info["optimal_value"], info["expected_return"], info["variance"], info["training time"], str(info["optimal_weights"])]
    # "Vectorized" = "Yes" hardcoded for now since it's always vectorized
    if not values_exist(sheet, values):
        row = find_first_empty_row(sheet)
        for i, value in enumerate(values, 1):
            sheet.cell(row, i, value)

    workbook.save(xlsx_name)


def save_to_xlsx_gurobi(cardinality_constraint, sigma, info, name_dataset, filename):
    """
    Saves the results to an Excel file.

    Attributes:
        cardinality_constraint (int): the cardinality constraint
        sigma (float): the sigma for the trade-off risk
        info (dict): the information about the training and testing times, and values
        name_dataset (str): the name of the dataset
        filename (str): the name of the Excel file to save the results to
    """
    try:
        workbook = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        print("File not found. Creating a new one.")
        workbook = openpyxl.Workbook()
    if "New" not in workbook.sheetnames:
        sheet = workbook.create_sheet("New")
        sheet.append(["Model", "Data", "Cardinality Constraint", "Sigma", "Optimal Value", "Expected Return", "Variance", "Solution Time", "Gap", "Optimal Weights", "Status"])
    else:
        sheet = workbook["New"]

    def find_first_empty_row(sheet):
        for row in range(2, 100000):
            if sheet[f"A{row}"].value is None:
                return row   
    
    values = ["Gurobi", name_dataset, cardinality_constraint, sigma, info["optimal_value"], info["expected_return"], info["variance"], info["solution_time"], info["gap"], str(info["optimal_weights"]), info["status"]]

    row = find_first_empty_row(sheet)
    for i, value in enumerate(values, 1):
        sheet.cell(row, i, value)

    workbook.save(filename)
#=============================================================================================


#=============================================================================================
#=================================Additional Functions========================================
#=============================================================================================
def read_stock_file(file_path):
    """
    Reads the stock file and returns the data as a pandas DataFrame.
    The stock files of this function are in the folder "Data".

    Attributes:
        file_path (str): the path to the file
    
    Returns:
        pd.DataFrame: the data
    """
    with open(file_path, 'r') as file:
        lines = file.readlines()
    
    N, T = map(int, lines[0].strip().split())
    all_values = []
    
    for line in lines[1:]:
        values = list(map(float, line.strip().split()))
        all_values.extend(values)
    
    expected_number_of_values = (N) * (T + 1)
    if len(all_values) != expected_number_of_values:
        N = len(all_values) // (T + 1)
    
    data = {f'Stock_{i+1}': all_values[i*(T+1):(i+1)*(T+1)] for i in range(N)}
    df = pd.DataFrame(data)
    
    return df



def random_weights(cardinality_constraint: int, num_shares: int, seed: int = 0):
    """
    Generate a weight vector with respect to the cardinality constraint.

    Attributes:
        cardinality_constraint (int): the cardinality constraint
        num_shares (int): the number of shares
        seed (int): the seed to use for the random generation, default is 0
    """
    np.random.seed(seed)
    weights = np.zeros(num_shares)
    random_index = np.random.choice(num_shares, cardinality_constraint, replace=False)
    weights[random_index] = np.random.rand(cardinality_constraint)
    weights /= weights.sum()
    return weights

    